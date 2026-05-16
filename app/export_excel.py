"""Client-ready multi-sheet Excel export (overview, monthly, yearly, charts, children, details)."""

from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.calculations import (
    chart_payload,
    children_total,
    compute_full_summary,
    monatlich,
    section_total,
    summe,
    total_1,
    total_2,
    total_3,
    total_4,
    total_5,
    total_revenue,
)
from app.categories import MONTH_NAMES_DE, MONTH_NAMES_EN, SECTIONS, SECTION_BY_ID, sections_for_tab
from app.future_invest import enrich_future_investment

_INCOME_SECTION_IDS = frozenset({
    "income",
    "self_employed_a",
    "self_employed_b",
    "net_a",
    "net_b",
    "other_income_a",
    "other_income_b",
})

_SUMMARY_BREAKDOWN_SECTIONS: dict[str, list[str]] = {
    "Total revenue": [s for s in _INCOME_SECTION_IDS if s in SECTION_BY_ID],
    "Total - 1 Living": ["living"],
    "Total - 2 Housing": ["housing"],
    "Total - 3 Insurance": ["health_a", "health_b", "property_insurance"],
    "Total - 4 Savings/Loans": ["pension", "wealth", "credit"],
    "Children (school & fees)": ["child_1", "child_2"],
}

# Styles
_FILL_HEADER = PatternFill("solid", fgColor="2563EB")
_FILL_SECTION = PatternFill("solid", fgColor="DBEAFE")
_FILL_SUBHEADER = PatternFill("solid", fgColor="E2E8F0")
_FILL_POSITIVE = PatternFill("solid", fgColor="D1FAE5")
_FILL_NEGATIVE = PatternFill("solid", fgColor="FEE2E2")
_FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
_FONT_TITLE = Font(bold=True, size=14)
_FONT_BOLD = Font(bold=True)
_FONT_SECTION = Font(bold=True, size=11)
_ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")
_THIN = Side(style="thin", color="CBD5E1")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_NUM_FMT = '#,##0.00'


def _style_header_row(ws, row: int, ncol: int) -> None:
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_row(ws, row: int, values: list, bold: bool = False, fmt_money: bool = False) -> None:
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = _BORDER
        if bold:
            cell.font = _FONT_BOLD
        if fmt_money and isinstance(val, (int, float)) and col > 1:
            cell.number_format = _NUM_FMT


def _chart_png(render_fn, *args, **kwargs) -> bytes | None:
    try:
        from app.export_pdf_charts import render_bar_png, render_donut_png, render_items_pie_png
    except ImportError:
        return None
    fns = {
        "donut": render_donut_png,
        "items": render_items_pie_png,
        "bar": render_bar_png,
    }
    fn = fns.get(render_fn)
    if not fn:
        return None
    return fn(*args, **kwargs)


def _embed_chart_png(
    ws,
    png: bytes | None,
    anchor: str,
    width: int = 480,
    height: int = 270,
) -> None:
    if not png:
        return
    img = XLImage(BytesIO(png))
    img.width = width
    img.height = height
    ws.add_image(img, anchor)


def _expense_item_slices(
    entries: dict[str, list[float]],
    custom_lines_by_section: dict[str, list[dict]] | None,
    max_visible: int = 12,
) -> list[dict]:
    """Top expense lines with amounts (annual summe), for category pie."""
    cl_map = custom_lines_by_section or {}
    items: list[dict] = []
    seen: set[str] = set()
    for section in SECTIONS:
        if section.id in _INCOME_SECTION_IDS:
            continue
        line_defs: list[tuple[str, str, str, str]] = [
            (line.key, line.label_en, line.label_de, line.line_type)
            for line in section.lines
        ]
        for cl in cl_map.get(section.id, []):
            line_defs.append(
                (
                    cl["line_key"],
                    cl.get("label_en") or cl["label_de"],
                    cl["label_de"],
                    cl.get("line_type", "expense"),
                )
            )
        for key, label_en, label_de, line_type in line_defs:
            if line_type == "income":
                continue
            if key in seen:
                continue
            vals = entries.get(key, [0.0] * 12)
            amount = summe(vals)
            if amount <= 0:
                continue
            seen.add(key)
            items.append({
                "id": key,
                "label": label_en,
                "label_en": label_en,
                "label_de": label_de,
                "amount": amount,
            })
    items.sort(key=lambda x: -x["amount"])
    if len(items) > max_visible:
        top = items[: max_visible - 1]
        other_amt = sum(i["amount"] for i in items[max_visible - 1 :])
        top.append({
            "id": "other",
            "label": "Other",
            "label_en": "Other",
            "label_de": "Sonstiges",
            "amount": other_amt,
        })
        items = top
    total = sum(i["amount"] for i in items) or 1
    for i in items:
        i["pct"] = round(100 * i["amount"] / total, 1)
    return items


def _top_expense_lines(
    entries: dict[str, list[float]],
    custom_lines_by_section: dict[str, list[dict]] | None,
    limit: int = 15,
) -> list[tuple[str, str, str, float]]:
    """(section_en, label_en, label_de, annual_total) sorted by amount."""
    cl_map = custom_lines_by_section or {}
    rows: list[tuple[str, str, str, float]] = []
    for section in SECTIONS:
        if section.id in _INCOME_SECTION_IDS:
            continue
        for line in section.lines:
            if line.line_type == "income":
                continue
            amt = summe(entries.get(line.key, [0.0] * 12))
            if amt > 0:
                rows.append((section.title_en, line.label_en, line.label_de, amt))
        for cl in cl_map.get(section.id, []):
            if cl.get("line_type") == "income":
                continue
            key = cl["line_key"]
            amt = summe(entries.get(key, [0.0] * 12))
            if amt > 0:
                rows.append((
                    section.title_en,
                    cl.get("label_en") or cl["label_de"],
                    cl["label_de"],
                    amt,
                ))
    rows.sort(key=lambda x: -x[3])
    return rows[:limit]


def _autosize(ws, max_col: int, min_width: int = 10, max_width: int = 42) -> None:
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        best = min_width
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value is not None:
                    best = max(best, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[letter].width = best


def _sheet_overview(wb: Workbook, client: dict, year: int, summary: dict, charts: dict) -> None:
    ws = wb.create_sheet("Overview", 0)
    ws["A1"] = "Annual Budget Report"
    ws["A1"].font = _FONT_TITLE
    ws.merge_cells("A1:D1")

    info = [
        ("Client", client.get("name", "")),
        ("Company", client.get("company_name") or "—"),
        ("Year", year),
        ("Report date", date.today().isoformat()),
    ]
    r = 3
    for label, val in info:
        ws.cell(row=r, column=1, value=label).font = _FONT_BOLD
        ws.cell(row=r, column=2, value=val)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Annual key figures (full year)").font = _FONT_SECTION
    r += 1
    _write_row(ws, r, ["Metric", "Amount", "Description"], bold=True)
    _style_header_row(ws, r, 3)
    r += 1

    rev = summary["total_revenue"]["summe"]
    exp = summary["total_expenses"]["summe"]
    diff = summary["difference"]["summe"]
    ch = summary.get("total_children", {}).get("summe", 0)

    bs = charts.get("balance_sheet") or {}
    net = bs.get("net_worth")
    kpis = [
        ("Total revenue", rev, "All income sources combined"),
        ("Total expenses", exp, "Living + housing + insurance + savings + children"),
        ("Balance (surplus / deficit)", diff, "Revenue minus expenses"),
        ("Children (school & fees)", ch, "Child 1 + Child 2 education costs"),
    ]
    if net is not None:
        kpis.extend([
            ("Total assets", bs.get("total_assets", 0), "Savings, gold, property, investments"),
            ("Total debts", bs.get("total_debts", 0), "Loans and liabilities outstanding"),
            ("Net worth", net, "Total assets minus total debts"),
        ])
    for label, amount, desc in kpis:
        _write_row(ws, r, [label, amount, desc], fmt_money=True)
        if amount < 0 and "Balance" in label:
            ws.cell(row=r, column=2).fill = _FILL_NEGATIVE
        elif "revenue" in label.lower() or (amount > 0 and "Balance" in label):
            if "Balance" in label and amount > 0:
                ws.cell(row=r, column=2).fill = _FILL_POSITIVE
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Expense breakdown (selected month average)").font = _FONT_SECTION
    r += 1
    _write_row(ws, r, ["Category", "Share %", "Amount"], bold=True)
    _style_header_row(ws, r, 3)
    r += 1
    for sl in charts.get("donut_sections", []):
        _write_row(
            ws,
            r,
            [f"{sl.get('label_de', sl.get('label', ''))} / {sl.get('label', '')}", sl.get("pct", 0), sl.get("amount", 0)],
            fmt_money=True,
        )
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="How to read this workbook").font = _FONT_SECTION
    r += 1
    notes = [
        "• Dashboard — KPI snapshot with chart pictures (PNG) for client meetings.",
        "• Monthly — month-by-month revenue, expenses, and balance.",
        "• Yearly — annual summary rows (same as app Summary tab).",
        "• Summary breakdown — only budget lines you entered (non-zero amounts).",
        "• Income — all income sources line by line.",
        "• Charts — tables, Excel charts, and picture charts (spending split + bar chart).",
        "• Children — school fees, bus, books, semester fees for Child 1 & Child 2.",
        "• Assets & Debts — assets, loans, interest, payoff plan; net worth.",
        "• Investment plan — goals, savings, targets (if recorded in the app).",
        "• Details — every budget line (including custom rows) with monthly amounts.",
    ]
    for note in notes:
        ws.cell(row=r, column=1, value=note)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1

    _autosize(ws, 4)


def _sheet_monthly(
    wb: Workbook,
    entries: dict,
    monatlich_mode: str,
    custom_keys_by_section: dict[str, list[str]] | None = None,
) -> None:
    ck = custom_keys_by_section
    ws = wb.create_sheet("Monthly")
    headers = ["Row"] + MONTH_NAMES_EN + ["Summe", "Avg/month"]
    _write_row(ws, 1, headers, bold=True)
    _style_header_row(ws, 1, len(headers))

    rows_def = [
        ("Total revenue", lambda m: total_revenue(entries, m, ck)),
        ("Total expenses", lambda m: total_5(entries, m, ck)),
        ("Children (school & fees)", lambda m: children_total(entries, m, ck)),
        ("Living", lambda m: total_1(entries, m, ck)),
        ("Housing", lambda m: total_2(entries, m, ck)),
        ("Insurance & health", lambda m: total_3(entries, m, ck)),
        ("Savings, pension & loans", lambda m: total_4(entries, m, ck)),
        (
            "Balance (revenue − expenses)",
            lambda m: total_revenue(entries, m, ck) - total_5(entries, m, ck),
        ),
    ]

    r = 2
    for label, fn in rows_def:
        months = [round(fn(m), 2) for m in range(1, 13)]
        _write_row(ws, r, [label] + months + [summe(months), monatlich(months, monatlich_mode)], fmt_money=True)
        if "Balance" in label:
            for c in range(2, 16):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, (int, float)) and v < 0:
                    ws.cell(row=r, column=c).fill = _FILL_NEGATIVE
                elif isinstance(v, (int, float)) and v > 0:
                    ws.cell(row=r, column=c).fill = _FILL_POSITIVE
        r += 1

    _autosize(ws, len(headers))


def _sheet_yearly(wb: Workbook, summary: dict) -> None:
    ws = wb.create_sheet("Yearly")
    headers = ["Summary row", "Description (DE)"] + MONTH_NAMES_EN + ["Summe", "Monatlich"]
    _write_row(ws, 1, headers, bold=True)
    _style_header_row(ws, 1, len(headers))

    row_defs = [
        ("Total revenue", "Gesamteinnahmen", summary["total_revenue"]),
        ("Total - 1 Living", "Lebenshaltung", summary["total_1"]),
        ("Total - 2 Housing", "Wohnen", summary["total_2"]),
        ("Total - 3 Insurance", "Versicherung & Gesundheit", summary["total_3"]),
        ("Total - 4 Savings/Loans", "Rente, Vermögen, Kredite", summary["total_4"]),
        ("Children (school & fees)", "Kinder — Schule & Gebühren", summary.get("total_children", {})),
        ("Total expenses", "Gesamtausgaben", summary["total_expenses"]),
        ("Difference (balance)", "Saldo / Differenz", summary["difference"]),
    ]

    r = 2
    for en, de, data in row_defs:
        if not data:
            data = {"months": [0] * 12, "summe": 0, "monatlich": 0}
        _write_row(
            ws,
            r,
            [en, de] + data.get("months", [0] * 12) + [data.get("summe", 0), data.get("monatlich", 0)],
            fmt_money=True,
        )
        if "Difference" in en:
            for c in range(3, 17):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, (int, float)) and v < 0:
                    ws.cell(row=r, column=c).fill = _FILL_NEGATIVE
                elif isinstance(v, (int, float)) and v > 0:
                    ws.cell(row=r, column=c).fill = _FILL_POSITIVE
        r += 1

    _autosize(ws, len(headers))


def _sheet_dashboard(
    wb: Workbook,
    client: dict,
    year: int,
    summary: dict,
    charts: dict,
    item_slices: list[dict],
) -> None:
    ws = wb.create_sheet("Dashboard")
    ws["A1"] = f"Budget dashboard — {client.get('name', '')} ({year})"
    ws["A1"].font = _FONT_TITLE
    ws.merge_cells("A1:F1")

    hero = charts.get("hero") or {}
    yr = charts.get("year") or {}
    r = 3
    _write_row(ws, r, ["Metric", "This month", "Full year"], bold=True)
    _style_header_row(ws, r, 3)
    r += 1
    dash_rows = [
        ("Revenue", hero.get("revenue", 0), yr.get("revenue", 0)),
        ("Expenses", hero.get("expenses", 0), yr.get("expenses", 0)),
        ("Balance", hero.get("difference", 0), yr.get("difference", 0)),
    ]
    for label, mo, annual in dash_rows:
        _write_row(ws, r, [label, mo, annual], fmt_money=True)
        r += 1

    bs = charts.get("balance_sheet") or {}
    if bs:
        r += 1
        _write_row(ws, r, ["Net worth", bs.get("net_worth", 0), ""], bold=True, fmt_money=True)
        r += 1

    r += 2
    section_slices = charts.get("donut_sections") or []
    if section_slices:
        _embed_chart_png(
            ws,
            _chart_png("donut", section_slices, "Spending by section (annual)"),
            f"A{r}",
        )
        r += 16
    if item_slices:
        _embed_chart_png(
            ws,
            _chart_png("items", item_slices, "Spending by category (annual)"),
            f"A{r}",
        )
        r += 16
    bars = charts.get("monthly_bars") or []
    if bars:
        _embed_chart_png(
            ws,
            _chart_png("bar", bars, "Revenue vs expenses by month"),
            f"A{r}",
            width=520,
            height=300,
        )

    _autosize(ws, 4)


def _sheet_charts(
    wb: Workbook,
    charts: dict,
    item_slices: list[dict],
    top_expenses: list[tuple[str, str, str, float]],
) -> None:
    ws = wb.create_sheet("Charts")
    ws["A1"] = "Charts & pictures (client-ready)"
    ws["A1"].font = _FONT_TITLE

    r = 3
    ws.cell(row=r, column=1, value="Spending by section (full year)").font = _FONT_SECTION
    r += 1
    _write_row(ws, r, ["Category (EN)", "Category (DE)", "Annual amount", "Share %"], bold=True)
    _style_header_row(ws, r, 4)
    r += 1
    section_slices = charts.get("donut_sections", [])
    sec_data_start = r
    for sl in section_slices:
        _write_row(
            ws,
            r,
            [sl.get("label", ""), sl.get("label_de", ""), sl.get("amount", 0), sl.get("pct", 0)],
            fmt_money=True,
        )
        r += 1

    if section_slices:
        pie = PieChart()
        pie.title = "Spending by section"
        pie.height = 9
        pie.width = 12
        data_ref = Reference(ws, min_col=3, min_row=sec_data_start - 1, max_row=r - 1)
        cats_ref = Reference(ws, min_col=1, min_row=sec_data_start, max_row=r - 1)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(cats_ref)
        ws.add_chart(pie, "F4")
        _embed_chart_png(
            ws,
            _chart_png("donut", section_slices, "Spending by section"),
            "F18",
            width=420,
            height=240,
        )

    r += 2
    ws.cell(row=r, column=1, value="Spending by category (filled lines, full year)").font = _FONT_SECTION
    r += 1
    _write_row(ws, r, ["Item (EN)", "Item (DE)", "Annual amount", "Share %"], bold=True)
    _style_header_row(ws, r, 4)
    r += 1
    item_start = r
    for sl in item_slices:
        _write_row(
            ws,
            r,
            [sl.get("label", ""), sl.get("label_de", ""), sl.get("amount", 0), sl.get("pct", 0)],
            fmt_money=True,
        )
        r += 1

    if item_slices:
        pie2 = PieChart()
        pie2.title = "By category"
        pie2.height = 9
        pie2.width = 12
        data_ref = Reference(ws, min_col=3, min_row=item_start - 1, max_row=r - 1)
        cats_ref = Reference(ws, min_col=1, min_row=item_start, max_row=r - 1)
        pie2.add_data(data_ref, titles_from_data=True)
        pie2.set_categories(cats_ref)
        ws.add_chart(pie2, f"F{r - len(item_slices) - 2}")
        _embed_chart_png(
            ws,
            _chart_png("items", item_slices, "Spending by category"),
            f"F{r + 2}",
            width=420,
            height=240,
        )
        r += 16

    r += 2
    ws.cell(row=r, column=1, value="Revenue vs expenses by month").font = _FONT_SECTION
    r += 1
    _write_row(ws, r, ["Month", "Revenue", "Expenses", "Balance"], bold=True)
    _style_header_row(ws, r, 4)
    r += 1
    bars = charts.get("monthly_bars", [])
    bar_start = r
    for b in bars:
        mo = int(b.get("month", 1))
        name = MONTH_NAMES_EN[mo - 1] if 1 <= mo <= 12 else str(mo)
        _write_row(
            ws,
            r,
            [name, b.get("revenue", 0), b.get("expenses", 0), b.get("difference", 0)],
            fmt_money=True,
        )
        r += 1

    if bars:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Revenue vs expenses"
        chart.y_axis.title = "Amount"
        chart.height = 11
        chart.width = 16
        data_ref = Reference(
            ws, min_col=2, min_row=bar_start - 1, max_row=bar_start - 1 + len(bars), max_col=3
        )
        cats_ref = Reference(ws, min_col=1, min_row=bar_start, max_row=bar_start - 1 + len(bars))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, f"F{bar_start}")
        _embed_chart_png(
            ws,
            _chart_png("bar", bars, "Revenue vs expenses"),
            f"A{r + 2}",
            width=520,
            height=300,
        )
        r += 18

    r += 2
    ws.cell(row=r, column=1, value="Top expenses (full year total)").font = _FONT_SECTION
    r += 1
    _write_row(ws, r, ["Section", "Item (EN)", "Item (DE)", "Annual amount"], bold=True)
    _style_header_row(ws, r, 4)
    r += 1
    for sec_en, en, de, amt in top_expenses:
        _write_row(ws, r, [sec_en, en, de, amt], fmt_money=True)
        r += 1

    _autosize(ws, 4)


def _sheet_children(
    wb: Workbook,
    entries: dict,
    monatlich_mode: str,
    custom_keys_by_section: dict[str, list[str]] | None = None,
) -> None:
    ck = custom_keys_by_section
    ws = wb.create_sheet("Children")
    ws["A1"] = "Children — school fees, books, bus, semester & yearly fees"
    ws["A1"].font = _FONT_TITLE
    ws.merge_cells("A1:O1")

    sections = sections_for_tab("children")
    r = 3
    headers = ["Expense item (DE)", "Expense item (EN)"] + MONTH_NAMES_EN + ["Summe", "Monatlich"]
    for section in sections:
        ws.cell(row=r, column=1, value=f"{section.title_de} / {section.title_en}").font = _FONT_SECTION
        ws.cell(row=r, column=1).fill = _FILL_SECTION
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
        r += 1
        _write_row(ws, r, headers, bold=True)
        _style_header_row(ws, r, len(headers))
        r += 1
        for line in section.lines:
            vals = entries.get(line.key, [0.0] * 12)
            _write_row(
                ws,
                r,
                [line.label_de, line.label_en] + vals + [summe(vals), monatlich(vals, monatlich_mode)],
                fmt_money=True,
            )
            r += 1
        if section.summary_key:
            months = [
                round(section_total(entries, section.id, m, ck.get(section.id) if ck else None), 2)
                for m in range(1, 13)
            ]
            ws.cell(row=r, column=1, value=section.summary_label_de or "Total").font = _FONT_BOLD
            ws.cell(row=r, column=2, value=section.summary_label_en or "Total").font = _FONT_BOLD
            for i, v in enumerate(months, start=3):
                c = ws.cell(row=r, column=i, value=v)
                c.number_format = _NUM_FMT
                c.font = _FONT_BOLD
            ws.cell(row=r, column=15, value=summe(months)).number_format = _NUM_FMT
            ws.cell(row=r, column=16, value=monatlich(months, monatlich_mode)).number_format = _NUM_FMT
            for c in range(1, 17):
                ws.cell(row=r, column=c).fill = _FILL_SUBHEADER
            r += 2

    _autosize(ws, len(headers))


def _sheet_details(
    wb: Workbook,
    entries: dict,
    monatlich_mode: str,
    custom_lines_by_section: dict[str, list[dict]] | None = None,
) -> None:
    ws = wb.create_sheet("Details")
    headers = ["Section", "Item (DE)", "Item (EN)"] + MONTH_NAMES_EN + ["Summe", "Monatlich"]
    _write_row(ws, 1, headers, bold=True)
    _style_header_row(ws, 1, len(headers))
    cl_map = custom_lines_by_section or {}
    r = 2
    for section in SECTIONS:
        ws.cell(row=r, column=1, value=f"{section.title_de} / {section.title_en}")
        ws.cell(row=r, column=1).fill = _FILL_SECTION
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
        r += 1
        for line in section.lines:
            vals = entries.get(line.key, [0.0] * 12)
            _write_row(
                ws,
                r,
                [section.title_en, line.label_de, line.label_en]
                + vals
                + [summe(vals), monatlich(vals, monatlich_mode)],
                fmt_money=True,
            )
            r += 1
        for cl in cl_map.get(section.id, []):
            key = cl["line_key"]
            vals = entries.get(key, [0.0] * 12)
            _write_row(
                ws,
                r,
                [
                    section.title_en,
                    f"{cl['label_de']} (custom)",
                    f"{cl.get('label_en', cl['label_de'])} (custom)",
                ]
                + vals
                + [summe(vals), monatlich(vals, monatlich_mode)],
                fmt_money=True,
            )
            r += 1
        r += 1
    _autosize(ws, len(headers))


def export_sheet_names() -> list[str]:
    """Expected workbook tabs for tests and documentation."""
    return [
        "Overview",
        "Dashboard",
        "Assets & Debts",
        "Investment plan",
        "Monthly",
        "Yearly",
        "Summary breakdown",
        "Income",
        "Charts",
        "Children",
        "Details",
    ]


def _sheet_income(
    wb: Workbook,
    entries: dict,
    monatlich_mode: str,
    custom_lines_by_section: dict[str, list[dict]] | None = None,
) -> None:
    ws = wb.create_sheet("Income")
    ws["A1"] = "Income — all sources (monthly amounts)"
    ws["A1"].font = _FONT_TITLE
    ws.merge_cells("A1:O1")
    headers = ["Section", "Item (DE)", "Item (EN)"] + MONTH_NAMES_EN + ["Summe", "Monatlich"]
    cl_map = custom_lines_by_section or {}
    r = 3
    for section in SECTIONS:
        if section.id not in _INCOME_SECTION_IDS:
            continue
        ws.cell(row=r, column=1, value=f"{section.title_de} / {section.title_en}").font = _FONT_SECTION
        ws.cell(row=r, column=1).fill = _FILL_SECTION
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
        r += 1
        _write_row(ws, r, headers, bold=True)
        _style_header_row(ws, r, len(headers))
        r += 1
        for line in section.lines:
            if line.line_type != "income":
                continue
            vals = entries.get(line.key, [0.0] * 12)
            _write_row(
                ws,
                r,
                [section.title_en, line.label_de, line.label_en]
                + vals
                + [summe(vals), monatlich(vals, monatlich_mode)],
                fmt_money=True,
            )
            r += 1
        for cl in cl_map.get(section.id, []):
            if cl.get("line_type") == "expense":
                continue
            key = cl["line_key"]
            vals = entries.get(key, [0.0] * 12)
            _write_row(
                ws,
                r,
                [
                    section.title_en,
                    f"{cl['label_de']} (custom)",
                    f"{cl.get('label_en', cl['label_de'])} (custom)",
                ]
                + vals
                + [summe(vals), monatlich(vals, monatlich_mode)],
                fmt_money=True,
            )
            r += 1
        r += 1
    _autosize(ws, len(headers))


def _sheet_summary_breakdown(
    wb: Workbook,
    entries: dict,
    monatlich_mode: str,
    custom_lines_by_section: dict[str, list[dict]] | None = None,
) -> None:
    """Only lines with non-zero amounts — matches app Summary expand rows."""
    ws = wb.create_sheet("Summary breakdown")
    ws["A1"] = "Summary breakdown — entered amounts only"
    ws["A1"].font = _FONT_TITLE
    ws.merge_cells("A1:O1")
    headers = ["Summary total", "Section", "Item (DE)", "Item (EN)"] + MONTH_NAMES_EN + ["Summe", "Monatlich"]
    cl_map = custom_lines_by_section or {}
    r = 3
    for total_label, section_ids in _SUMMARY_BREAKDOWN_SECTIONS.items():
        ws.cell(row=r, column=1, value=total_label).font = _FONT_SECTION
        ws.cell(row=r, column=1).fill = _FILL_SECTION
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
        r += 1
        any_line = False
        for sid in section_ids:
            section = SECTION_BY_ID.get(sid)
            if not section:
                continue
            line_defs: list[tuple] = [(line, False) for line in section.lines]
            for cl in cl_map.get(sid, []):
                line_defs.append((cl, True))
            for item, is_custom in line_defs:
                if is_custom:
                    key = item["line_key"]
                    label_de = item["label_de"]
                    label_en = item.get("label_en") or label_de
                    line_type = item.get("line_type", "expense")
                else:
                    key = item.key
                    label_de = item.label_de
                    label_en = item.label_en
                    line_type = item.line_type
                if total_label == "Total revenue" and line_type != "income":
                    continue
                if total_label != "Total revenue" and line_type == "income":
                    continue
                vals = entries.get(key, [0.0] * 12)
                if not any(v != 0 for v in vals):
                    continue
                any_line = True
                _write_row(
                    ws,
                    r,
                    [total_label, section.title_en, label_de, label_en]
                    + vals
                    + [summe(vals), monatlich(vals, monatlich_mode)],
                    fmt_money=True,
                )
                r += 1
        if not any_line:
            ws.cell(row=r, column=1, value="(no amounts entered)")
            r += 1
        r += 1
    _autosize(ws, len(headers))


def _sheet_investments(wb: Workbook, investments: list[dict], year: int) -> None:
    ws = wb.create_sheet("Investment plan")
    ws["A1"] = "Investment goals & projections"
    ws["A1"].font = _FONT_TITLE
    r = 3
    headers = [
        "Name",
        "Type",
        "Already invested",
        "Target",
        "Monthly savings",
        "Return % p.a.",
        "Target year",
        "Progress %",
        "Gap",
        "Months to target",
        "Projected value",
        "Notes",
    ]
    _write_row(ws, r, headers, bold=True)
    _style_header_row(ws, r, len(headers))
    r += 1
    for row in investments:
        enriched = enrich_future_investment(row, year)
        _write_row(
            ws,
            r,
            [
                enriched.get("name", ""),
                enriched.get("investment_type", ""),
                enriched.get("current_amount", 0),
                enriched.get("target_amount", 0),
                enriched.get("monthly_contribution", 0),
                enriched.get("expected_return_annual", 0),
                enriched.get("target_year", ""),
                enriched.get("progress_pct", ""),
                enriched.get("gap", ""),
                enriched.get("months_to_target", ""),
                enriched.get("projected_value", ""),
                enriched.get("notes", ""),
            ],
            fmt_money=True,
        )
        r += 1
    if not investments:
        ws.cell(row=r, column=1, value="No investment goals recorded in the app.")
    _autosize(ws, len(headers))


def _sheet_balance_sheet(wb: Workbook, balance: dict) -> None:
    ws = wb.create_sheet("Assets & Debts")
    ws["A1"] = "Assets & debts (current values)"
    ws["A1"].font = _FONT_TITLE
    r = 3
    payoff = balance.get("payoff") or {}
    if payoff:
        ws.cell(row=r, column=1, value="Debt payoff plan (from budget)").font = _FONT_SECTION
        r += 1
        plan_rows = [
            ("Monthly income", payoff.get("monthly_income", 0)),
            ("Monthly expenses", payoff.get("monthly_expenses", 0)),
            ("Monthly surplus", payoff.get("monthly_surplus", 0)),
            ("Repay / month (entered)", payoff.get("total_monthly_payment", 0)),
        ]
        for label, val in plan_rows:
            _write_row(ws, r, [label, val], fmt_money=True)
            r += 1
        months = payoff.get("months_to_clear_debts")
        if months is not None:
            _write_row(
                ws,
                r,
                ["Months to clear all debts", months, payoff.get("debt_free_label", "")],
            )
            r += 1
        r += 1

    ws.cell(row=r, column=1, value="Assets").font = _FONT_SECTION
    r += 1
    _write_row(ws, r, ["Name", "Type", "Current value", "Notes"], bold=True)
    _style_header_row(ws, r, 4)
    r += 1
    for a in balance.get("assets", []):
        _write_row(
            ws,
            r,
            [a.get("name", ""), a.get("asset_type", ""), a.get("amount", 0), a.get("notes", "")],
            fmt_money=True,
        )
        r += 1
    _write_row(ws, r, ["Total assets", "", balance.get("total_assets", 0), ""], bold=True, fmt_money=True)
    r += 2
    ws.cell(row=r, column=1, value="Debts").font = _FONT_SECTION
    r += 1
    _write_row(
        ws,
        r,
        ["Name", "Interest % p.a.", "Outstanding", "Repay / month", "Months to clear", "Notes"],
        bold=True,
    )
    _style_header_row(ws, r, 6)
    r += 1
    per = {p.get("id"): p for p in (payoff.get("per_debt") or [])}
    for d in balance.get("debts", []):
        pd = per.get(d.get("id")) or {}
        months = pd.get("months_to_clear")
        if months is None and pd.get("payment_covers_interest") is False:
            months_txt = "!"
        elif months is not None:
            months_txt = str(months)
        else:
            months_txt = ""
        _write_row(
            ws,
            r,
            [
                d.get("name", ""),
                d.get("interest_rate_annual", 0),
                d.get("amount", 0),
                d.get("monthly_payment", 0),
                months_txt,
                d.get("notes", ""),
            ],
            fmt_money=True,
        )
        r += 1
    _write_row(
        ws,
        r,
        [
            "Total debts",
            "",
            balance.get("total_debts", 0),
            payoff.get("total_monthly_payment", 0),
            "",
            "",
        ],
        bold=True,
        fmt_money=True,
    )
    r += 2
    _write_row(ws, r, ["Net worth (assets − debts)", balance.get("net_worth", 0)], bold=True, fmt_money=True)
    _autosize(ws, 6)


def _annual_donut_slices(entries: dict, summary: dict) -> list[dict]:
    from app.calculations import baufinanzierung_total, build_monthly_series

    slices = [
        {"label": "Living", "label_de": "Lebenshaltung", "amount": summary["total_1"]["summe"]},
        {"label": "Housing", "label_de": "Wohnen", "amount": summary["total_2"]["summe"]},
        {"label": "Insurance", "label_de": "Versicherung", "amount": summary["total_3"]["summe"]},
        {"label": "Savings & Loans", "label_de": "Sparen & Kredite", "amount": summary["total_4"]["summe"]},
    ]
    bau = summe(build_monthly_series(entries, baufinanzierung_total))
    if bau > 0:
        slices.append({"label": "Financing", "label_de": "Baufinanzierung", "amount": bau})
    ch = summary.get("total_children", {}).get("summe", 0)
    if ch > 0:
        slices.append({"label": "Children", "label_de": "Kinder", "amount": ch})
    total = sum(s["amount"] for s in slices) or 1
    for s in slices:
        s["pct"] = round(100 * s["amount"] / total, 1)
    return slices


def build_client_workbook(
    client: dict,
    year: int,
    entries: dict[str, list[float]],
    monatlich_mode: str,
    custom_keys_by_section: dict[str, list[str]] | None = None,
    balance_sheet: dict | None = None,
    custom_lines_by_section: dict[str, list[dict]] | None = None,
    future_investments: list[dict] | None = None,
) -> BytesIO:
    ck = custom_keys_by_section
    cl = custom_lines_by_section
    summary = compute_full_summary(entries, monatlich_mode, ck)
    charts = chart_payload(entries, 8, monatlich_mode, ck)
    charts["donut_sections"] = _annual_donut_slices(entries, summary)
    if balance_sheet:
        charts["balance_sheet"] = balance_sheet

    item_slices = _expense_item_slices(entries, cl)
    top_expenses = _top_expense_lines(entries, cl)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    _sheet_overview(wb, client, year, summary, charts)
    _sheet_dashboard(wb, client, year, summary, charts, item_slices)
    if balance_sheet:
        _sheet_balance_sheet(wb, balance_sheet)
    _sheet_investments(wb, future_investments or [], year)
    _sheet_monthly(wb, entries, monatlich_mode, ck)
    _sheet_yearly(wb, summary)
    _sheet_summary_breakdown(wb, entries, monatlich_mode, cl)
    _sheet_income(wb, entries, monatlich_mode, cl)
    _sheet_charts(wb, charts, item_slices, top_expenses)
    _sheet_children(wb, entries, monatlich_mode, ck)
    _sheet_details(wb, entries, monatlich_mode, cl)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
