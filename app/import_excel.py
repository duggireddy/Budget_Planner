"""Excel import utilities."""

from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.categories import LINE_BY_KEY, SECTIONS
from app.database import get_mappings, save_mapping


def parse_de_amount(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).replace("€", "").strip()
    if not s:
        return 0.0
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _normalize_label(label: str) -> str:
    return re.sub(r"\s+", " ", label.strip().lower())


def match_line_key(label: str, mappings: dict[str, str]) -> str | None:
    if label in mappings:
        return mappings[label]
    norm = _normalize_label(label)
    for file_label, key in mappings.items():
        if _normalize_label(file_label) == norm:
            return key
    for section in SECTIONS:
        for line in section.lines:
            if norm in (_normalize_label(line.label_de), _normalize_label(line.label_en)):
                return line.key
            if _normalize_label(line.label_en) in norm or _normalize_label(line.label_de) in norm:
                return line.key
    return None


MONTH_HEADERS = {
    "januar", "january", "jan", "februar", "february", "feb",
    "märz", "marz", "march", "mar", "april", "apr", "mai", "may",
    "juni", "june", "jun", "juli", "july", "jul", "august", "aug",
    "september", "sep", "oktober", "october", "oct", "november", "nov",
    "dezember", "december", "dec",
}


def parse_workbook(content: bytes) -> tuple[list[dict], list[str]]:
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], ["Empty workbook"]

    header_idx = None
    month_cols: list[tuple[int, int]] = []
    for i, row in enumerate(rows[:20]):
        labels = [str(c).strip().lower() if c else "" for c in row]
        for j, lab in enumerate(labels):
            if lab in MONTH_HEADERS:
                if header_idx is None:
                    header_idx = i
                break
        if header_idx is not None:
            for j, lab in enumerate(labels):
                lab_clean = lab.replace(".", "")
                month_num = _month_number(lab_clean)
                if month_num:
                    month_cols.append((j, month_num))
            break

    if header_idx is None or not month_cols:
        return [], ["Could not find month headers (Januar, Feb, …)"]

    month_cols.sort(key=lambda x: x[1])
    mappings = get_mappings()
    parsed: list[dict] = []
    warnings: list[str] = []

    for row in rows[header_idx + 1 :]:
        if not row or not row[0]:
            continue
        label = str(row[0]).strip()
        if not label or label.lower().startswith("summe") and "total" not in label.lower():
            if "total revenue" in label.lower() or "total -" in label.lower():
                continue
        if label.startswith("Summe") or label.startswith("Total -"):
            continue

        key = match_line_key(label, mappings)
        status = "matched" if key else "unmapped"
        amounts: dict[int, float] = {}
        for col_idx, month_num in month_cols:
            if col_idx < len(row):
                amounts[month_num] = parse_de_amount(row[col_idx])

        if any(amounts.values()):
            parsed.append({
                "label": label,
                "line_key": key,
                "status": status,
                "amounts": amounts,
            })
        if status == "unmapped" and any(amounts.values()):
            warnings.append(f"Unmapped row: {label}")

    return parsed, warnings


def _month_number(header: str) -> int | None:
    mapping = {
        "januar": 1, "january": 1, "jan": 1,
        "februar": 2, "february": 2, "feb": 2,
        "märz": 3, "marz": 3, "march": 3, "mar": 3,
        "april": 4, "apr": 4,
        "mai": 5, "may": 5,
        "juni": 6, "june": 6, "jun": 6,
        "juli": 7, "july": 7, "jul": 7,
        "august": 8, "aug": 8,
        "september": 9, "sep": 9,
        "oktober": 10, "october": 10, "oct": 10,
        "november": 11, "nov": 11,
        "dezember": 12, "december": 12, "dec": 12,
    }
    return mapping.get(header.lower())


def apply_import(budget_id: int, rows: list[dict]) -> int:
    from app.database import upsert_entry

    count = 0
    for row in rows:
        key = row.get("line_key")
        if not key or key not in LINE_BY_KEY:
            continue
        for month, amount in row.get("amounts", {}).items():
            upsert_entry(budget_id, key, month, amount)
            count += 1
    return count


def all_category_options() -> list[dict]:
    opts = []
    for section in SECTIONS:
        for line in section.lines:
            opts.append({
                "key": line.key,
                "label": f"{section.title_en} / {line.label_en}",
                "section": section.id,
            })
    return opts
