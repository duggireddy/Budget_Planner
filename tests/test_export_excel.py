"""Excel export workbook structure and content."""

from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.database import (
    create_asset,
    create_custom_line,
    create_debt,
    custom_line_keys_by_section,
    custom_lines_by_section,
    get_or_create_budget,
    upsert_entry,
)
from app.export_excel import build_client_workbook, export_sheet_names
from app.seed import ensure_sample_client


@pytest.fixture
def sample_export(api_client):
    cid = ensure_sample_client()
    bid = get_or_create_budget(cid, 2026)
    line = create_custom_line(bid, "living", "Gym", "Gym", "expense")
    upsert_entry(bid, line["line_key"], 3, 75.0)
    create_asset(cid, {"name": "Cash", "asset_type": "savings", "amount": 2000})
    create_debt(cid, {"name": "Credit card", "amount": 500})
    from app.database import compute_balance_sheet, get_client, get_settings, load_entries

    client = get_client(cid)
    entries = load_entries(bid)
    settings = get_settings(bid)
    ck = custom_line_keys_by_section(bid)
    balance = compute_balance_sheet(cid)
    cl = custom_lines_by_section(bid)
    buf = build_client_workbook(
        client, 2026, entries, settings["monatlich_mode"], ck, balance, cl
    )
    return load_workbook(BytesIO(buf.getvalue()), read_only=True, data_only=True)


def test_export_has_all_sheets(sample_export):
    names = sample_export.sheetnames
    for expected in export_sheet_names():
        assert expected in names


def test_balance_sheet_net_worth_row(sample_export):
    ws = sample_export["Assets & Debts"]
    found = False
    for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
        if row and row[0] and "Net worth" in str(row[0]):
            assert row[1] == 1500.0
            found = True
            break
    assert found


def test_details_includes_custom_line(sample_export):
    ws = sample_export["Details"]
    labels = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) > 2 and row[2]:
            labels.append(str(row[2]))
    assert any("custom" in lb.lower() for lb in labels)
    assert any("Gym" in lb for lb in labels)


def test_monthly_has_expense_rows(sample_export):
    ws = sample_export["Monthly"]
    row_labels = [ws.cell(row=i, column=1).value for i in range(2, 12)]
    assert "Total expenses" in row_labels
    assert "Children (school & fees)" in row_labels
